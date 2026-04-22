import openpyxl
import numpy as np
import scipy.stats as stats

def get_heights_pairwise(frame, conversion_function=lambda x: x):
    male_female_pairs = []

    for row in range(2, frame.max_row):
        pair = []
        for col in frame.iter_cols(1, frame.max_column):
            pair.append(conversion_function(col[row].value))
        male_female_pairs.append(pair)

    return male_female_pairs

def get_heights(dataframe, column, conversion_function=lambda x: x):
    heights = []
    for row in range(2, dataframe.max_row):
        heights.append(conversion_function(dataframe[row][column].value))
    return heights

def get_z_score(x, mu, sigma):
    return x - mu / sigma

marsh_data = openpyxl.load_workbook("heights-data-marsh-1988-great-britain.xlsx")
marsh_sheet = marsh_data.active

galton_data = openpyxl.load_workbook("GaltonFamilies-1886.xlsx")
galton_sheet = galton_data.active

male_column = 0
female_column = 1

marsh_male_heights = get_heights(marsh_sheet, male_column)
marsh_female_heights = get_heights(marsh_sheet, female_column)

galton_male_heights = get_heights(galton_sheet, male_column, lambda x: 25.4*x)
galton_female_heights = get_heights(galton_sheet, female_column, lambda x: 25.4*x)

marsh_male_heights_mean = np.mean(marsh_male_heights)
marsh_female_heights_mean = np.mean(marsh_female_heights)
galton_male_heights_mean = np.mean(galton_male_heights)
galton_female_heights_mean = np.mean(galton_female_heights)

marsh_male_heights_std = np.std(marsh_male_heights)
marsh_female_heights_std = np.std(marsh_female_heights)
galton_male_heights_std = np.std(galton_male_heights)
galton_female_heights_std = np.std(galton_female_heights)


# Mean
print("SAMPLE MEANS")
print(f"Marsh Mean male height: {marsh_male_heights_mean:.3f} mm")
print(f"Marsh Mean female height: {marsh_female_heights_mean:.3f} mm")
print(f"Galton Mean male height: {galton_male_heights_mean:.3f} mm")
print(f"Galton Mean female height: {galton_female_heights_mean:.3f} mm")
print()

# Standard deviation
print("SAMPLE STANDARD DEVIATIONS")
print(f"Marsh male height standard deviation: {marsh_male_heights_std:.3f} mm")
print(f"Marsh female height standard deviation: {marsh_female_heights_std:.3f} mm")
print(f"Galton male height standard deviation: {galton_male_heights_std:.3f} mm")
print(f"Galton female height standard deviation: {galton_female_heights_std:.3f} mm")
print()

# correlation

marsh_correlation = np.corrcoef(marsh_male_heights, marsh_female_heights)[0, 1]
galton_correlation = np.corrcoef(marsh_male_heights, marsh_female_heights)[0, 1]

print("SAMPLE MEAN CORRELATION")
print(f"Correlation between male and female heights (Marsh) {marsh_correlation:.3f}")
print(f"Correlation between male and female heights (Galton) {galton_correlation:.3f}")
print()

# differences
print("THEORETICAL DIFFERENCE MEAN AND STANDARD DEVIATIONS")
print(f"marsh E(x1-x2) = {marsh_male_heights_mean - marsh_female_heights_mean:.3f} mm")
print(f"galton E(x1-x2) = {galton_male_heights_mean - galton_female_heights_mean:.3f} mm")
marsh_var_of_difference = marsh_male_heights_std**2 + marsh_female_heights_std**2 - 2*marsh_correlation*marsh_male_heights_std*marsh_female_heights_std
print(f"marsh Var(x1-x2) = {marsh_var_of_difference:.3f} mm^2")
print(f"\tstd of that is {np.sqrt(marsh_var_of_difference):.3f} mm")
galton_var_of_difference = galton_male_heights_std**2 + galton_female_heights_std**2 - 2*galton_correlation*galton_male_heights_std*galton_male_heights_std
print(f"galton Var(x1-x2) = {galton_var_of_difference:.3f} mm^2")
print(f"\tstd of that is {np.sqrt(galton_var_of_difference):.3f} mm")
marsh_theoretical_z_score_of_difference = get_z_score(0, marsh_male_heights_mean - marsh_female_heights_mean, np.sqrt(marsh_var_of_difference))
print(f"(marsh) Z score of difference=0: {marsh_theoretical_z_score_of_difference:.3f}")
print(f"P(z > {marsh_theoretical_z_score_of_difference:.3f}): {stats.norm.sf(marsh_theoretical_z_score_of_difference):.3f}")
galton_theoretical_z_score_of_difference = get_z_score(0, galton_male_heights_mean - galton_female_heights_mean, np.sqrt(galton_var_of_difference))
print(f"(galton) z score of difference=0: {galton_theoretical_z_score_of_difference:.3f}")
print(f"P(z > {galton_theoretical_z_score_of_difference:.3f}): {stats.norm.sf(galton_theoretical_z_score_of_difference):.3f}")
print()

# height difference calculations
marsh_pairs = get_heights_pairwise(marsh_sheet)
galton_pairs = get_heights_pairwise(galton_sheet, lambda x: 25.4*x)

marsh_height_differences = []
for pair in marsh_pairs:
    marsh_height_differences.append(pair[0] - pair[1])
marsh_height_differences = np.array(marsh_height_differences)

galton_height_differences = []
for pair in galton_pairs:
    galton_height_differences.append(pair[0] - pair[1])
galton_height_differences = np.array(galton_height_differences)

print("SAMPLE DIFFERENCE MEAN AND STANDARD DEVIATIONS")
print(f"Marsh height difference mean: {np.mean(marsh_height_differences):.3f} mm")
print(f"Marsh height difference std: {np.std(marsh_height_differences):.3f} mm")
print(f"Galton height difference mean: {np.mean(galton_height_differences):.3f} mm")
print(f"Galton height difference std: {np.std(galton_height_differences):.3f} mm")
print()