import openpyxl
import numpy as np

def get_heights_pairwise(frame):
    male_female_pairs = []

    for row in range(2, frame.max_row):
        pair = []
        for col in frame.iter_cols(1, frame.max_column):
            pair.append(col[row].value)
        male_female_pairs.append(pair)

    return male_female_pairs

def get_heights(dataframe, column, conversion_function=lambda x: x):
    heights = []
    for row in range(2, dataframe.max_row):
        heights.append(conversion_function(dataframe[row][column].value))
    return heights



marsh_data = openpyxl.load_workbook("heights-data-marsh-1988-great-britain.xlsx")
marsh_active_sheet = marsh_data.active

galton_data = openpyxl.load_workbook("GaltonFamilies-1886.xlsx")
galton_sheet = galton_data.active

male_column = 0
female_column = 1

marsh_male_heights = get_heights(marsh_active_sheet, male_column)
marsh_female_heights = get_heights(marsh_active_sheet, female_column)

galton_male_heights = get_heights(galton_sheet, male_column, lambda x: 25.4*x)
galton_female_heights = get_heights(galton_sheet, female_column, lambda x: 25.4*x)

marsh_male_heights_mean = np.mean(marsh_male_heights)
marsh_female_heights_mean = np.mean(marsh_female_heights)
galton_male_heights_mean = np.mean(galton_male_heights)
galton_female_heights_mean = np.mean(galton_female_heights)


# Mean
print(f"Marsh Mean male height: {marsh_male_heights_mean:.3f} mm")
print(f"Marsh Mean female height: {marsh_female_heights_mean:.3f} mm")
print(f"Galton Mean male height: {galton_male_heights_mean:.3f} mm")
print(f"Galton Mean female height: {galton_female_heights_mean:.3f} mm")
print(f"marsh x1-x2 = {marsh_male_heights_mean - marsh_female_heights_mean:.3f} mm")
print(f"galton x1-x2 = {galton_male_heights_mean - galton_female_heights_mean:.3f} mm")

# Standard deviation
print()
print(f"Marsh male height standard deviation: {np.std(marsh_male_heights):.3f} mm")
print(f"Marsh female height standard deviation: {np.std(marsh_female_heights):.3f} mm")
print(f"Galton male height standard deviation: {np.std(galton_male_heights):.3f} mm")
print(f"Galton female height standard deviation: {np.std(galton_female_heights):.3f} mm")

# correlation
print()
print(f"Correlation between male and female heights (Marsh) {np.corrcoef(marsh_male_heights, marsh_female_heights)[0, 1]:.3f}")
print(f"Correlation between male and female heights (Galton) {np.corrcoef(galton_male_heights, galton_female_heights)[0, 1]:.3f}")